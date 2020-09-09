# Python Core
import os.path

# Excel Library
import openpyxl
from openpyxl import load_workbook


class Controller:
    """
    Class that handles excel books
    """

    def __init__(self):
        self.wb = openpyxl.Workbook()   # Inicialize Workbook
        self.filename = 'money_control.xlsx'
        self.sheet_name = 'Control de Gastos'


    def check_excel_exist(self):
        excel_exist = os.path.isfile(self.filename)

        if excel_exist:
            return True     # Excel Exist
        else:
            return False    # Excel doesn't Exist


    def add_data_to_excel(self, initial_balance, data):
        exist = self.check_excel_exist()
        if exist:
            wb = load_workbook(filename=self.filename)
            sheet_expense_control = wb[self.sheet_name]

            # Add Expenses or Income
            sheet_expense_control.append(data)          # <------ This not found. Developing

        else:
            sheet = self.wb.active              # Active Sheet
            sheet.title = self.sheet_name       # Rename Sheet


            initial_data = [
                ('Saldo inicial:', initial_balance),    # Initial Balance
                (' ',),                                 # Blank Space
                ('Ingresos', 'Gastos', 'Fecha'),     # Sheet Header
            ]

            # Add Initial Data
            for tuple_data in initial_data:
                sheet.append(tuple_data)

            # Add Expenses or Income
            sheet.append(data)

            # Save Excel
            self.wb.save(self.filename)